"""UNP Grading System - xlsx gradesheet exporter.

Writes a workbook in the E-Class Record style with LIVE formulas, so
the sheet itself shows the computation and recalculates in Excel:
INPUT DATA (roster + transmutation table), Midterm, Finals, and Final
Semestral Grade sheets.
"""
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

import engine

THIN = Side(style="thin", color="9AA39C")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HDR_FILL = PatternFill("solid", fgColor="E3EFE6")
HPS_FILL = PatternFill("solid", fgColor="FFF6D9")
CALC_FILL = PatternFill("solid", fgColor="EFF3EF")
FONT = "Calibri"


def _cell(ws, row, col, value=None, bold=False, center=False, fill=None,
          border=True, size=10, number_format=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(name=FONT, size=size, bold=bold)
    if center:
        c.alignment = Alignment(horizontal="center", vertical="center",
                                wrap_text=True)
    else:
        c.alignment = Alignment(vertical="center", wrap_text=True)
    if fill:
        c.fill = fill
    if border:
        c.border = BORDER
    if number_format:
        c.number_format = number_format
    return c


def _term_sheet(wb, cls, term_key, title, exam_label, first_student_row=13):
    """Build one term sheet. Returns dict of layout info for cross refs."""
    cfg = cls["config"]
    ws = wb.create_sheet(title)
    m = cls["meta"]

    _cell(ws, 1, 1, "UNIVERSITY OF NORTHERN PHILIPPINES", bold=True,
          border=False, size=12)
    _cell(ws, 2, 1, "College of Communication and Information Technology",
          border=False)
    _cell(ws, 3, 1, title.upper() + " GRADING SHEET", bold=True,
          border=False, size=11)
    _cell(ws, 5, 1, f"Subject: {m.get('subject_code','')} - "
                    f"{m.get('subject_title','')}", border=False)
    _cell(ws, 5, 10, f"Section: {m.get('section','')}", border=False)
    _cell(ws, 5, 16, f"School Year: {m.get('school_year','')}",
          border=False)
    _cell(ws, 6, 1, f"Instructor: {m.get('instructor','')}", border=False)
    _cell(ws, 6, 10, f"Term: {m.get('term','')}", border=False)
    _cell(ws, 6, 16, f"Schedule: {m.get('schedule','')}", border=False)

    hdr1, hdr2, hps_row = 9, 10, 11
    col = 1
    _cell(ws, hdr1, col, "No", bold=True, center=True, fill=HDR_FILL)
    ws.merge_cells(start_row=hdr1, start_column=1, end_row=hdr2,
                   end_column=1)
    col += 1
    _cell(ws, hdr1, col, "LEARNERS' NAMES", bold=True, center=True,
          fill=HDR_FILL)
    ws.merge_cells(start_row=hdr1, start_column=2, end_row=hdr2,
                   end_column=2)
    ws.column_dimensions["B"].width = 34
    col += 1

    layout = {}   # comp -> dict(score_cols, total, ps, ws)
    for comp in engine.COMPONENTS:
        n = cfg["columns"][comp]
        label = cfg["labels"][comp]
        if comp == "exam":
            label = exam_label
        weight = cfg["weights"][comp]
        start = col
        _cell(ws, hdr1, start, f"{label} ({weight}%)", bold=True,
              center=True, fill=HDR_FILL)
        ws.merge_cells(start_row=hdr1, start_column=start, end_row=hdr1,
                       end_column=start + n + 2)
        for i in range(n):
            _cell(ws, hdr2, col, i + 1, bold=True, center=True,
                  fill=HDR_FILL)
            ws.column_dimensions[get_column_letter(col)].width = 6
            col += 1
        t_col, ps_col, ws_col = col, col + 1, col + 2
        for c, t in ((t_col, "Total"), (ps_col, "PS"), (ws_col, "WS")):
            _cell(ws, hdr2, c, t, bold=True, center=True, fill=HDR_FILL)
            ws.column_dimensions[get_column_letter(c)].width = 7
        col += 3
        layout[comp] = {"scores": list(range(start, start + n)),
                        "total": t_col, "ps": ps_col, "ws": ws_col}

    init_col, grade_col, remark_col = col, col + 1, col + 2
    for c, t in ((init_col, "INITIAL\nGRADE"), (grade_col, "GRADE"),
                 (remark_col, "REMARK")):
        _cell(ws, hdr1, c, t, bold=True, center=True, fill=HDR_FILL)
        ws.merge_cells(start_row=hdr1, start_column=c, end_row=hdr2,
                       end_column=c)
        ws.column_dimensions[get_column_letter(c)].width = 9
    layout["result"] = {"initial": init_col, "grade": grade_col,
                        "remark": remark_col}

    # HPS row
    _cell(ws, hps_row, 2, "HIGHEST POSSIBLE SCORE", bold=True,
          fill=HPS_FILL)
    _cell(ws, hps_row, 1, "", fill=HPS_FILL)
    for comp in engine.COMPONENTS:
        info = layout[comp]
        for i, c in enumerate(info["scores"]):
            hps = cfg["hps"][term_key][comp]
            v = hps[i] if i < len(hps) else None
            _cell(ws, hps_row, c, v, center=True, fill=HPS_FILL)
        sc = [get_column_letter(c) for c in info["scores"]]
        rng = f"{sc[0]}{hps_row}:{sc[-1]}{hps_row}"
        _cell(ws, hps_row, info["total"],
              f'=IF(COUNT({rng})=0,"",SUM({rng}))', center=True,
              fill=HPS_FILL)
        _cell(ws, hps_row, info["ps"], 100, center=True, fill=HPS_FILL)
        _cell(ws, hps_row, info["ws"], cfg["weights"][comp] / 100.0,
              center=True, fill=HPS_FILL)

    # student rows
    students = cls["students"]
    for idx, s in enumerate(students):
        r = first_student_row + idx
        _cell(ws, r, 1, idx + 1, center=True)
        _cell(ws, r, 2, f"='INPUT DATA'!B{11 + idx}")
        for comp in engine.COMPONENTS:
            info = layout[comp]
            vals = cls["scores"].get(term_key, {}).get(s["id"], {}).get(
                comp, [])
            for i, c in enumerate(info["scores"]):
                v = vals[i] if i < len(vals) else None
                _cell(ws, r, c, v, center=True)
            sc = [get_column_letter(c) for c in info["scores"]]
            rng = f"{sc[0]}{r}:{sc[-1]}{r}"
            tL = get_column_letter(info["total"])
            pL = get_column_letter(info["ps"])
            wL = get_column_letter(info["ws"])
            _cell(ws, r, info["total"],
                  f'=IF(COUNT({rng})=0,"",SUM({rng}))', center=True,
                  fill=CALC_FILL)
            _cell(ws, r, info["ps"],
                  f'=IFERROR(IF({tL}{r}="","",ROUND({tL}{r}/${tL}$'
                  f'{hps_row}*50+50,2)),"")', center=True, fill=CALC_FILL)
            _cell(ws, r, info["ws"],
                  f'=IFERROR(IF({pL}{r}="","",ROUND({pL}{r}*${wL}$'
                  f'{hps_row},2)),"")', center=True, fill=CALC_FILL)
        wsum = "+".join(f"{get_column_letter(layout[c]['ws'])}{r}"
                        for c in engine.COMPONENTS)
        iL = get_column_letter(init_col)
        _cell(ws, r, init_col, f'=IFERROR({wsum},"")', center=True,
              fill=CALC_FILL, number_format="0.00")
        _cell(ws, r, grade_col,
              f"=IFERROR(VLOOKUP({iL}{r},'INPUT DATA'!$AL$11:$AN$24,3),"
              f'"")', center=True, fill=CALC_FILL, number_format="0.00")
        _cell(ws, r, remark_col,
              f'=IF({iL}{r}="","",IF({iL}{r}>=75,"PASSED","FAILED"))',
              center=True, fill=CALC_FILL)

    ws.freeze_panes = ws.cell(row=first_student_row, column=3)
    return {"sheet": title, "first_row": first_student_row,
            "initial_col": get_column_letter(init_col),
            "grade_col": get_column_letter(grade_col)}


def export_gradesheet(cls, path, include=("midterm", "finals",
                                          "semestral")):
    """Write the gradesheet workbook. include controls which parts."""
    cfg = cls["config"]
    m = cls["meta"]
    wb = Workbook()

    # ---------------- INPUT DATA ----------------
    ws = wb.active
    ws.title = "INPUT DATA"
    _cell(ws, 1, 1, "Input Data Sheet for E-Class Record", bold=True,
          border=False, size=12)
    info = [("SECTION:", m.get("section", "")),
            ("SCHOOL YEAR:", m.get("school_year", "")),
            ("TERM:", m.get("term", "")),
            ("SUBJECT CODE:", m.get("subject_code", "")),
            ("SUBJECT DESCRIPTION:", m.get("subject_title", "")),
            ("COURSE:", m.get("course", cls["students"][0]["course"]
                              if cls["students"] else "")),
            ("INSTRUCTOR:", m.get("instructor", "")),
            ("SCHEDULE:", m.get("schedule", ""))]
    for i, (k, v) in enumerate(info):
        _cell(ws, 3 + i, 4, k, bold=True, border=False)
        _cell(ws, 3 + i, 7, v, border=False)

    _cell(ws, 10, 1, "No", bold=True, center=True, fill=HDR_FILL)
    _cell(ws, 10, 2, "LEARNERS' NAMES", bold=True, fill=HDR_FILL)
    ws.column_dimensions["B"].width = 34
    for i, s in enumerate(cls["students"]):
        _cell(ws, 11 + i, 1, i + 1, center=True)
        _cell(ws, 11 + i, 2, s["name"])

    # transmutation table at AL11:AN24 (same address his sheets use)
    _cell(ws, 10, 38, "From", bold=True, center=True, fill=HDR_FILL)
    _cell(ws, 10, 39, "To", bold=True, center=True, fill=HDR_FILL)
    _cell(ws, 10, 40, "Grade", bold=True, center=True, fill=HDR_FILL)
    _cell(ws, 10, 41, "Descriptor", bold=True, fill=HDR_FILL)
    table = cfg["transmutation"]
    for i, (bound, grade, desc) in enumerate(table):
        upper = (table[i + 1][0] - 0.01) if i + 1 < len(table) else 100
        _cell(ws, 11 + i, 38, bound, center=True)          # AL
        _cell(ws, 11 + i, 39, upper, center=True)          # AM
        _cell(ws, 11 + i, 40, grade, center=True,          # AN
              number_format="0.00")
        _cell(ws, 11 + i, 41, desc)                        # AO

    refs = {}
    if "midterm" in include:
        refs["midterm"] = _term_sheet(wb, cls, "midterm", "Midterm",
                                      "MIDTERM EXAM")
    if "finals" in include:
        refs["finals"] = _term_sheet(wb, cls, "finals", "Finals",
                                     "FINAL EXAM")

    # ---------------- Semestral ----------------
    if "semestral" in include and "midterm" in refs and "finals" in refs:
        ws = wb.create_sheet("Final Semestral Grade")
        _cell(ws, 1, 1, "UNIVERSITY OF NORTHERN PHILIPPINES", bold=True,
              border=False, size=12)
        _cell(ws, 2, 1, "FINAL SEMESTRAL GRADE", bold=True, border=False)
        _cell(ws, 4, 1, f"Subject: {m.get('subject_code','')} - "
                        f"{m.get('subject_title','')}", border=False)
        _cell(ws, 4, 8, f"Section: {m.get('section','')}", border=False)
        _cell(ws, 4, 12, f"SY: {m.get('school_year','')} "
                         f"{m.get('term','')}", border=False)
        mw = cfg["semestral_weights"]["midterm"]
        fw = cfg["semestral_weights"]["finals"]
        _cell(ws, 6, 4, mw, center=True, fill=HPS_FILL)   # D6
        _cell(ws, 6, 6, fw, center=True, fill=HPS_FILL)   # F6
        _cell(ws, 6, 3, "Weight:", bold=True, border=False)
        hdr = ["No", "LEARNERS' NAMES", "MIDTERM\nINITIAL", "MIDTERM\nGRADE",
               "FINALS\nINITIAL", "FINALS\nGRADE", "SEMESTRAL\nINITIAL",
               "SEMESTRAL\nGRADE", "REMARK"]
        for c, t in enumerate(hdr, start=1):
            _cell(ws, 8, c, t, bold=True, center=True, fill=HDR_FILL)
        ws.column_dimensions["B"].width = 34
        for c in "CDEFGHI":
            ws.column_dimensions[c].width = 11
        mid, fin = refs["midterm"], refs["finals"]
        for i, s in enumerate(cls["students"]):
            r = 9 + i
            src = mid["first_row"] + i
            _cell(ws, r, 1, i + 1, center=True)
            _cell(ws, r, 2, f"='INPUT DATA'!B{11 + i}")
            _cell(ws, r, 3, f"=Midterm!{mid['initial_col']}{src}",
                  center=True, number_format="0.00")
            _cell(ws, r, 4, f"=Midterm!{mid['grade_col']}{src}",
                  center=True, number_format="0.00")
            _cell(ws, r, 5, f"=Finals!{fin['initial_col']}{src}",
                  center=True, number_format="0.00")
            _cell(ws, r, 6, f"=Finals!{fin['grade_col']}{src}",
                  center=True, number_format="0.00")
            _cell(ws, r, 7,
                  f'=IFERROR(IF(OR(C{r}="",E{r}=""),"",'
                  f'ROUND(C{r}*$D$6+E{r}*$F$6,2)),"")',
                  center=True, fill=CALC_FILL, number_format="0.00")
            _cell(ws, r, 8,
                  f"=IFERROR(VLOOKUP(G{r},'INPUT DATA'!$AL$11:$AN$24,3),"
                  f'"")', center=True, fill=CALC_FILL,
                  number_format="0.00")
            _cell(ws, r, 9,
                  f'=IF(G{r}="","",IF(G{r}>=75,"PASSED","FAILED"))',
                  center=True, fill=CALC_FILL)
        ws.freeze_panes = "C9"

    wb.save(path)
    return path
